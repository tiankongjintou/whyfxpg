"""
Excel report renderer: ReportRenderer port implementation using openpyxl.
"""
from pathlib import Path
from typing import Any

import openpyxl

from whyfxpg.ports.report_renderer import ReportRenderer
from whyfxpg.services.report_model import ReportModel


class ExcelReportRenderer(ReportRenderer):
    """Render a ReportModel as an .xlsx workbook with multiple sheets."""

    def __init__(self, db_path: str | None = None):
        # db_path is kept for API compatibility but no longer used;
        # all data comes from the ReportModel seam.
        self._db_path = db_path

    def render(self, report_model: ReportModel, output_path: Path) -> Path:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()

        # 风险事件明细
        ws = wb.active
        ws.title = "风险事件明细"
        self._append_dicts(ws, report_model.top_events)

        # 产品汇总
        ws2 = wb.create_sheet("产品风险汇总")
        self._append_dicts(ws2, report_model.top_products)

        # 预警清单
        ws3 = wb.create_sheet("预警清单")
        self._append_dicts(ws3, report_model.pending_alerts)

        # 执行摘要
        ws4 = wb.create_sheet("执行摘要")
        ws4.append(["生成时间", report_model.generated_at])
        ws4.append(["报告类型", report_model.report_type])
        ws4.append(["事件总数", report_model.total_events])
        ws4.append(["风险等级分布", str(report_model.level_counts)])
        ws4.append(["执行摘要"])
        ws4.append([report_model.executive_summary])

        wb.save(str(output_path))
        return output_path

    def _append_dicts(self, ws: Any, rows: list[dict[str, Any]]) -> None:
        if rows:
            headers = list(rows[0].keys())
            ws.append(headers)
            for row in rows:
                ws.append([row.get(h) for h in headers])
